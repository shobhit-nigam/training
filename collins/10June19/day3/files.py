#!/usr/bin/env python
# coding: utf-8

# In[3]:


import os
objf = open("hello.txt", 'w')


# In[4]:


objf.write("some line\n")


# In[5]:


objf.close()


# In[6]:


objf = open("hello.txt", 'w')


# In[7]:


objf.write("some line\n")
objf.write("some lines\n")
objf.write("some more lines\n")
objf.write("many more lines\n")


# In[8]:


objf.close()


# In[9]:


help(open)


# In[16]:


objr = open("hello.txt", 'r+')


# In[11]:


varx = objr.read(7)


# In[12]:


varx


# In[13]:


objr.write('new data\n')


# In[14]:


objr.close()


# In[17]:


varx = objr.read(7)


# In[18]:


varx = objr.read(7)


# In[19]:


varx


# In[20]:


print(varx)


# In[22]:


objr.tell()


# In[23]:


from openpyxl import load_workbookd_workbook


# In[24]:


wb = load_workbook("sample.xlsx")


# In[25]:


type(wb)


# In[26]:


print(wb.sheetnames)


# In[27]:


sheet = wb.get_sheet_by_name('imp_sheet')


# In[28]:


sheet = wb['imp_sheet']


# In[29]:


sheet.title


# In[31]:


sheet['B3'].value


# In[32]:


varc = sheet['B3']


# In[33]:


type(varc)


# In[34]:


varc.row


# In[35]:


varc.column


# In[36]:


varc.coordinate


# In[37]:


sheet.cell(row=2, column = 2).value


# In[38]:


for i in range(1,4):
    print(sheet.cell(row=i, column = 2).value)


# In[44]:


from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# In[40]:


file = "new_file.xlsx"


# In[45]:


wb = Workbook()


# In[47]:


wsa = wb.active


# In[48]:


wsa.title = "some_name"


# In[49]:


for row in range(1,20):
    wsa.append(range(300))


# In[50]:


wsb = wb.create_sheet(title = "second")


# In[51]:


wb.save(filename= file)


# In[52]:


wb = Workbook()


# In[53]:


wsa = wb.active


# In[54]:


wsa['A1'] = '20'


# In[55]:


wsa['B1'] = '30'


# In[56]:


wsa['C1'] = "=SUM(1,2)"


# In[57]:


wsa.merge_cells('A3:D3')


# In[58]:


from openpyxl.drawing.image import Image


# In[59]:


img = Image("d.png")


# In[60]:


wsa.add_image(img, 'F6')


# In[61]:


wb.save("all.xlsx")


# In[62]:


wb = Workbook()


# In[63]:


wsa = wb.active


# In[67]:


for i in range(10):
    wsa.append([i])


# In[69]:


from openpyxl.chart import BarChart, Reference, Series


# In[71]:


chart = BarChart()


# In[73]:


val = Reference(wsa, min_col=1, min_row=1, max_col=1, max_row=10)
chart.add_data(val)


# In[74]:


wsa.add_chart(chart, 'D10')


# In[75]:


wb.save("chart.xlsx")


# In[ ]:




